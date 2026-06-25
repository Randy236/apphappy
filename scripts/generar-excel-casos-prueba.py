# -*- coding: utf-8 -*-
"""Genera Happy Jump - Pruebas Funcionales Sistema.xlsx desde plantilla IEEE 829."""
import csv
import shutil
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font

ROOT = Path(__file__).resolve().parent.parent
TEMPLATE = Path(r"c:\Users\LIZBETH\Downloads\Plantilla_Pruebas_Funcionales_Sistema.xlsx")
FALLBACK_TEMPLATE = ROOT / "docs" / "plantillas" / "Plantilla_Pruebas_Funcionales_Sistema.xlsx"
CSV_CASES = ROOT / "docs" / "CASOS_DE_PRUEBA.csv"
OUT = ROOT / "docs" / "Happy_Jump_Pruebas_Funcionales_Sistema.xlsx"

PRIORITY = {"P0": "Alta", "P1": "Media", "P2": "Baja"}

DATOS_ENTRADA = {
    "CP-001": "Usuario: Rosisela\nPIN: 1234",
    "CP-002": "Usuario: Admin\nPIN: 1234",
    "CP-003": "Usuario: Admin\nPIN: 9999 (incorrecto)",
    "CP-004": "PIN: vacío o menos de 4 dígitos",
    "CP-005": "Mismo usuario logueado en otro dispositivo",
    "CP-006": "URL: http://IP_PC:3000/health",
    "CP-007": "Fecha: día con reservas de prueba",
    "CP-008": "Cliente, deporte Futbol/Voley, fecha futura, hora libre, montos",
    "CP-009": "Fecha u hora en el pasado",
    "CP-010": "Filtro opcional por salón",
    "CP-011": "Salón, horario libre, tipo evento, precio",
    "CP-012": "Periodo: diario / semanal / mensual",
    "CP-013": "PIN actual y PIN nuevo (mín. 4 dígitos)",
    "CP-014": "Build actual desplegado",
    "CP-015": "Mes con reservas en distintos días",
    "CP-016": "Calendario mensual abierto",
    "CP-017": "Barra de semana visible",
    "CP-018": "Reserva multi-franja mismo cliente",
    "CP-019": "Turno largo existente + motivo cancelación",
    "CP-020": "Hora inicio pocos minutos después de ahora",
    "CP-021": "Rango ej. 17:00 a 18:30",
    "CP-022": "Fecha distinta a hoy seleccionada",
}

EVIDENCIA = {
    "CP-001": "evidencias-pruebas/CP-001-login-trabajador.png",
    "CP-002": "evidencias-pruebas/CP-002-login-admin.png",
    "CP-006": "evidencias-pruebas/CP-006-health.png",
    "CP-007": "evidencias-pruebas/CP-007-lista-cancha.png",
    "CP-008": "evidencias-pruebas/CP-008-crear-reserva.png",
    "CP-015": "evidencias-pruebas/CP-015-calendario.png",
}


def fmt_pasos(pasos: str) -> str:
    parts = [p.strip() for p in pasos.split(") ") if p.strip()]
    out = []
    for i, p in enumerate(parts, 1):
        if not p.endswith(")"):
            p = p + ")" if ")" in pasos else p
        out.append(f"{i}. {p.lstrip('0123456789.) ')}")
    return "\n".join(out) if out else pasos.replace(") ", ")\n")


def load_cases():
    rows = []
    with CSV_CASES.open(encoding="utf-8", newline="") as f:
        for row in csv.DictReader(f):
            rows.append(row)
    return rows


def main():
    src = TEMPLATE if TEMPLATE.exists() else FALLBACK_TEMPLATE
    if not src.exists():
        print(f"No se encuentra plantilla en:\n  {TEMPLATE}\n  {FALLBACK_TEMPLATE}")
        return 1

    OUT.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, OUT)
    wb = openpyxl.load_workbook(OUT)
    cases = load_cases()
    total = len(cases)

    # --- Plan de Pruebas (valores en columna D; C está fusionada) ---
    plan = wb["Plan de Pruebas"]
    plan["D6"] = "Happy Jump (app móvil + API REST)"
    plan["D7"] = "1.0.0 / entrega final"
    plan["D8"] = "Login, Cancha, Salones, Reportes, Perfil"
    plan["D9"] = "Equipo Happy Jump"
    plan["D10"] = str(date.today())
    plan["D11"] = str(date.today())
    plan["D12"] = "Android físico/emulador; API Node :3000; MySQL (Laragon)"
    plan["D13"] = "Postman, Allure, Jenkins, k6, SonarCloud, Snyk, Excel"
    # Fila 17 usa fórmulas COUNTIF sobre columna Resultado — no sobrescribir

    # --- Casos de Prueba ---
    ws = wb["Casos de Prueba"]
    # Borrar filas de ejemplo (4 en adelante hasta fila 50)
    if ws.max_row >= 4:
        ws.delete_rows(4, ws.max_row - 3)

    header_row = 3
    cols = {ws.cell(header_row, c).value: c for c in range(1, 17)}

    wrap = Alignment(wrap_text=True, vertical="top")
    for i, c in enumerate(cases, start=4):
        cid = c["ID"]
        ws.cell(i, cols["ID"], cid)
        ws.cell(i, cols["Módulo"], c["Modulo"])
        ws.cell(i, cols["Nombre del Caso"], c["Nombre del caso"])
        ws.cell(i, cols["Tipo"], c["Tipo"])
        ws.cell(i, cols["Prioridad"], PRIORITY.get(c["Prioridad"], c["Prioridad"]))
        ws.cell(i, cols["Precondiciones"], c["Precondiciones"])
        ws.cell(i, cols["Datos de Entrada"], DATOS_ENTRADA.get(cid, "—"))
        ws.cell(i, cols["Pasos de Ejecución"], fmt_pasos(c["Pasos"]))
        ws.cell(i, cols["Resultado Esperado"], c["Resultado esperado"])
        ws.cell(i, cols["Resultado Obtenido"], None)
        ws.cell(i, cols["Evidencia"], EVIDENCIA.get(cid, ""))
        ws.cell(i, cols["Resultado"], "PENDIENTE")
        ws.cell(i, cols["Defecto ID"], None)
        ws.cell(i, cols["Ejecutado por"], None)
        ws.cell(i, cols["Fecha Ejecución"], None)
        ws.cell(i, cols["Observaciones"], "Ejecutar en celular + API activa")
        for col in range(1, 17):
            ws.cell(i, col).alignment = wrap
        ws.row_dimensions[i].height = 72

    ws.cell(1, 1, "CASOS DE PRUEBA — HAPPY JUMP (Funcionales y Sistema)")
    ws.cell(1, 1).font = Font(bold=True, size=14)

    # --- Defectos: limpiar ejemplo BUG-042 ---
    def_ws = wb["Defectos"]
    if def_ws.max_row >= 4:
        def_ws.delete_rows(4, def_ws.max_row - 3)

    # --- Métricas (opcional; celdas fusionadas varían) ---
    metrics_name = next((n for n in wb.sheetnames if "trica" in n.lower()), None)
    if metrics_name:
        try:
            m = wb[metrics_name]
            for addr, val in [("D6", "Happy Jump - ciclo entrega final"), ("D7", str(date.today()))]:
                if type(m[addr]).__name__ != "MergedCell":
                    m[addr] = val
        except Exception:
            pass

    wb.save(OUT)
    print(f"OK: {OUT}")
    print(f"Casos cargados: {total} ({cases[0]['ID']} … {cases[-1]['ID']})")
    print("Abre el Excel, ejecuta pruebas y cambia Resultado a PASS/FAIL.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
